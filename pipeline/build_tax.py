"""
UK Wealth Dashboard — tax-burden data pipeline.

Produces data/tax.json, which powers the "What share goes in tax?" panel.

Two kinds of input, kept strictly separate and each labelled with its source:

  1. MEASURED (ONS Effects of Taxes and Benefits, Table 8): total tax and its
     components as a % of gross income, by income quintile. This is the
     empirical "ordinary households" curve — and it shows the system is roughly
     FLAT (~30-37%) once regressive VAT/council tax is added to progressive
     income tax.

  2. CURATED constants, each citing a publication:
       - income-tax / NI bands 2025/26 (GOV.UK) — the frontend computes the
         user's direct tax from these.
       - the top-tail effective-rate curve from Advani, Hughson & Summers
         (2023): effective rates peak ~38% at £500k of remuneration then FALL
         below 30% above £3m.
       - a £1bn "poorest billionaire" benchmark with EXPLICIT, adjustable
         return/realisation assumptions — nothing hidden.

Run:  python pipeline/build_tax.py   (or `npm run data`, which runs both)
"""

from __future__ import annotations

import datetime as dt
import json
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
RAW = Path(__file__).resolve().parent / "raw"
RAW.mkdir(parents=True, exist_ok=True)

ETB_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
    "personalandhouseholdfinances/incomeandwealth/datasets/"
    "theeffectsoftaxesandbenefitsonhouseholdincomefinancialyearending2014/"
    "financialyearending2024/etbreferencetablesfye202324final.xlsx"
)
ETB_XLSX = RAW / "etb_fye2024.xlsx"

QUINTILE_LABELS = ["Bottom", "2nd", "3rd", "4th", "Top"]
QUINTILE_COLS = [3, 4, 5, 6, 7]  # columns in ETB Table 8


def fetch() -> None:
    if ETB_XLSX.exists():
        print(f"[tax] cached {ETB_XLSX.name}")
        return
    print(f"[tax] downloading {ETB_URL}")
    req = urllib.request.Request(ETB_URL, headers={"User-Agent": "Mozilla/5.0"})
    ETB_XLSX.write_bytes(urllib.request.urlopen(req, timeout=90).read())
    print(f"[tax] wrote {ETB_XLSX.stat().st_size} bytes")


def _row_by_label(rows, label: str) -> list[float]:
    """Return the quintile values for the first row whose label cell matches."""
    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(cell.startswith(label) for cell in cells[:3]):
            return [float(row[c]) for c in QUINTILE_COLS]
    raise ValueError(f"ETB row {label!r} not found")


def parse_etb() -> dict:
    wb = openpyxl.load_workbook(ETB_XLSX, read_only=True, data_only=True)
    t8 = list(wb["Table 8"].iter_rows(values_only=True))
    t2 = list(wb["Table 2a"].iter_rows(values_only=True))

    # Table 2a row: "Quintile points (equivalised £)" -> the 4 upper boundaries.
    boundaries: list[float] = []
    for row in t2:
        cells = [str(c) if c is not None else "" for c in row]
        if any(c.startswith("Quintile points") for c in cells):
            boundaries = [float(v) for v in row if isinstance(v, (int, float))]
            break
    if len(boundaries) < 4:
        raise ValueError("quintile boundaries not found")
    boundaries = boundaries[:4]

    return {
        "boundaries": boundaries,
        "allTaxes": _row_by_label(t8, "All taxes"),
        "incomeTax": _row_by_label(t8, "Income Tax"),
        "ni": _row_by_label(t8, "National Insurance"),
        "council": _row_by_label(t8, "Council Tax"),
        "indirect": _row_by_label(t8, "Indirect taxes"),
    }


def representative_incomes(boundaries: list[float]) -> list[int]:
    """A single income to plot each quintile at (approx midpoints; open-ended
    top quintile placed above the last boundary)."""
    b = boundaries
    reps = [
        b[0] * 0.55,
        (b[0] + b[1]) / 2,
        (b[1] + b[2]) / 2,
        (b[2] + b[3]) / 2,
        b[3] * 1.7,
    ]
    return [round(r / 100) * 100 for r in reps]


def build() -> None:
    fetch()
    etb = parse_etb()
    reps = representative_incomes(etb["boundaries"])

    tax = {
        "meta": {
            "generated": dt.date.today().isoformat(),
            "pipelineVersion": "1.0.0",
            "measuredSourceId": "ons_etb",
            "period": "FYE 2024 (2023/24)",
            "geography": "UK",
            "note": "Quintile rates are MEASURED (ONS ETB Table 8, % of gross "
                    "income). Income-tax bands, the top-tail curve and the "
                    "billionaire benchmark are curated — see sourceId on each.",
        },
        # MEASURED — ONS ETB Table 8, all households, % of gross income
        "quintiles": {
            "sourceId": "ons_etb",
            "label": QUINTILE_LABELS,
            "boundaryEquivalisedGbp": [round(x) for x in etb["boundaries"]],
            "representativeIncomeGbp": reps,
            "allTaxesPctGross": etb["allTaxes"],
            "incomeTaxPctGross": etb["incomeTax"],
            "niPctGross": etb["ni"],
            "councilPctGross": etb["council"],
            "indirectPctGross": etb["indirect"],
        },
        # CURATED — GOV.UK income-tax & NI rules, 2025/26 (England)
        "incomeTax2025_26": {
            "sourceId": "hmrc_income_tax_2025_26",
            "personalAllowance": 12570,
            "paTaperStart": 100000,     # £1 lost per £2 above this
            "paTaperEnd": 125140,       # allowance fully gone
            "bands": [
                {"threshold": 12570, "rate": 0.20},   # basic
                {"threshold": 50270, "rate": 0.40},   # higher
                {"threshold": 125140, "rate": 0.45},  # additional
            ],
            "ni": {  # employee Class 1, 2025/26
                "primaryThreshold": 12570,
                "upperEarningsLimit": 50270,
                "mainRate": 0.08,
                "upperRate": 0.02,
            },
        },
        # CURATED — Advani, Hughson & Summers (2023). Effective AVERAGE tax rate
        # on total remuneration (taxable income + realised gains), individuals.
        "topTailAdvani": {
            "sourceId": "advani_2023",
            "note": "Effective average tax rate on remuneration (income + "
                    "realised capital gains). Peaks ~38% at £500k then falls "
                    "below 30% above £3m; a quarter of those with £3m+ paid "
                    "12% or less.",
            "points": [
                [100000, 35], [250000, 37], [500000, 38],
                [1000000, 35], [3000000, 30], [5000000, 27], [10000000, 21],
            ],
            "quarterOf3mPlusPaidAtMostPct": 12,
        },
        # CURATED — "poorest billionaire" benchmark, built bottom-up so it has
        # the SAME rows as a normal bill. Every assumption explicit.
        "billionaire": {
            "sourceId": "advani_2023",
            "wealthGbp": 1_000_000_000,
            "homeEquityGbp": 30_000_000,      # a mansion; generates no income
            "economicReturnPct": 8,           # annual return on invested wealth
            "realisedTaxableYieldPct": 2,     # realised, taxable income as % invested
            "realisedGainsSharePct": 60,      # of realised income; the rest dividends
            "cgtRatePct": 24,                 # 2025/26 higher-rate CGT (GOV.UK)
            "dividendRatePct": 39.35,         # 2025/26 additional-rate dividend
            "salaryGbp": 0,                   # wealth, not wages -> no income tax/NI
            "annualSpendingGbp": 2_000_000,   # lifestyle spend subject to VAT
            "vatEffectivePct": 7.7,           # ONS top-quintile VAT as % of spend
            "councilTaxGbp": 4000,            # one top-band home (illustrative)
            "note": "Illustrative and adjustable. 8%/yr is conservative: UBS puts "
                    "billionaire wealth up 121% over the decade (~8%/yr, 10%/yr "
                    "in 2015-20) and Oxfam counts +$2tn (~15%) in 2024. Only ~2% "
                    "of wealth (~£20m) is realised as taxable income — 60% capital "
                    "gains at 24%, 40% dividends at 39.35%, a blended ~30% "
                    "matching Advani et al. The rest of the gain is unrealised, so "
                    "untaxed — raising the return only widens that untaxed slice.",
        },
        # CURATED — national tax take + an illustrative "fair share" scenario.
        "taxTake": {
            "sourceId": "obr_receipts",
            "currentReceiptsGbp": 1_139_000_000_000,      # 2024-25, OBR
            "top01WealthShare": 0.09,                     # WID (top 0.1%)
            "top01SourceId": "wid_uk",
            "totalHouseholdWealthGbp": 13_568_000_000_000,
            "note": "Illustrative. Most tax already comes from ordinary "
                    "households paying about a third of their income. The "
                    "shortfall is at the very top: this compares the tax the "
                    "wealthiest 0.1% pay on the income their wealth earns with "
                    "what they would pay at the median household's effective "
                    "rate. Uses this dashboard's own top-rate and return "
                    "assumptions, so treat it as illustrative, not a forecast.",
        },
        # CURATED — "what could the extra revenue pay for" unit costs.
        "payFor": {
            "taxCutPerPennyGbp": 6_900_000_000,   # HMRC: 1p on basic rate ≈ £6.9bn
            "taxCutSourceId": "hmrc_reckoner",
            "corpTaxPerPointGbp": 3_600_000_000,  # ≈ 1pp main rate (HMRC reckoner)
            "corpTaxRatePct": 25,
            "vatPerPointGbp": 8_500_000_000,      # ≈ 1pp standard rate (HMRC reckoner)
            "vatRatePct": 20,
            "stampDutyGbp": 10_400_000_000,       # residential SDLT, England 2024-25
            "stampDutySourceId": "hmrc_stamp",
            "councilTaxGbp": 41_000_000_000,      # council tax collected, England 2024-25
            "councilTaxSourceId": "council_tax",
            "nurseCostGbp": 45_000,               # salary + pension + employer NI ≈
            "doctorCostGbp": 100_000,
            "teacherCostGbp": 47_000,
            "staffSourceId": "gov_pay",
            "nhsNurses": 467_000,                 # nurses & midwives, all settings
            "nhsDoctors": 216_000,                # doctors, all settings (incl GPs)
            "teacherCount": 513_400,              # England, headcount 2024
            "workforceSourceId": "nhs_workforce",
            "teacherSourceId": "school_workforce",
            "freeUniversityGbp": 10_000_000_000,  # abolish tuition fees, IFS ~£8-11bn
            "freeUniversitySourceId": "ifs_tuition",
            "policeOfficers": 148_886,            # England & Wales, Sept 2024
            "policeCostGbp": 70_000,              # all-in cost of an extra officer ≈
            "policeSourceId": "police_workforce",
            "socialCareGbp": 9_000_000_000,       # free personal care, Health Foundation
            "socialCareSourceId": "health_foundation",
            "homeGrantGbp": 140_000,              # gov grant per social home (NHF/CEBR)
            "homeSourceId": "nhf_homes",
            "defenceBudgetGbp": 60_200_000_000,   # 2024-25, Commons Library
            "defenceSourceId": "commons_defence",
            "note": "Illustrative equivalents for the extra revenue. Staff "
                    "figures are approximate full employment cost (salary + "
                    "pension + employer NI); each option is the whole sum spent "
                    "one way, not all at once.",
        },
    }

    validate(tax)
    (DATA / "tax.json").write_text(
        json.dumps(tax, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    q = tax["quintiles"]
    print(f"[tax] quintile all-taxes % of gross: {q['allTaxesPctGross']}")
    print(f"[tax] boundaries £: {q['boundaryEquivalisedGbp']}  reps £: {q['representativeIncomeGbp']}")
    print("[tax] wrote data/tax.json")


def validate(tax: dict) -> None:
    q = tax["quintiles"]
    for key in ("allTaxesPctGross", "incomeTaxPctGross", "niPctGross",
                "councilPctGross", "indirectPctGross"):
        if len(q[key]) != 5:
            raise SystemExit(f"[tax][validate] FAIL: {key} must have 5 quintiles")
    lo, hi = q["allTaxesPctGross"][0], q["allTaxesPctGross"][-1]
    if not (20 <= lo <= 50 and 20 <= hi <= 50):
        raise SystemExit("[tax][validate] FAIL: all-taxes rates out of sane range")
    b = q["boundaryEquivalisedGbp"]
    if b != sorted(b):
        raise SystemExit("[tax][validate] FAIL: quintile boundaries not increasing")
    print("[tax][validate] OK")


if __name__ == "__main__":
    build()
